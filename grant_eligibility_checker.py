#!/usr/bin/env python3
"""
AlgoOrange Australian Grant Eligibility Analyzer (Groq edition)
Hybrid deterministic-blocker + LLM scoring with formatted Excel output.
"""

import argparse
import json
import logging
import os
import sys
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from groq import Groq
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
GROQ_MODEL = "openai/gpt-oss-120b"
GROQ_TEMPERATURE = 0.2
GROQ_MAX_TOKENS = 1200
REQUEST_TIMEOUT_S = 60
MAX_RETRIES = 4
RETRY_BASE_DELAY_S = 2.0
RATE_LIMIT_SLEEP_S = 1.0

ONLY_OPEN_GRANTS = True
INPUT_FILE_DEFAULT = "C:/Users/Manoj/OneDrive/Desktop/fall_detection/web_crawling/output_20260629_142108.json"
OUTPUT_DIR = "output"

PURSUE_THRESHOLD = 40
TIER_LIKELY = 70
TIER_POSSIBLE = 40

# ---------------------------------------------------------------------------
# AlgoOrange compliance profile
# ---------------------------------------------------------------------------
ORGANIZATION_INFO: Dict[str, Any] = {
    "name": "AlgoOrange",
    "tagline": "AI-Powered Vision Intelligence, Automation & Enterprise Software Solutions",
    "location": "Sydney, Australia",
    "country_primary": "Australia",
    "business_type": (
        "AI-powered software development and intelligent automation company specializing in "
        "Vision AI, IoT integration, workflow automation, compliance platforms, property-risk "
        "assessment (PropTech), and custom enterprise software."
    ),
    "team": {
        "size": 3,
        "profile": "Software engineers in AI, computer vision, LLMs, cloud, and system integration.",
    },
    "industry_domains": [
        "Artificial Intelligence", "Computer Vision", "Machine Learning",
        "Industrial Automation", "Workflow Automation", "Enterprise Software",
        "IoT", "SaaS", "PropTech / Property Risk Assessment",
        "Bushfire Risk (BAL / AS 3959)", "Flood Risk Screening",
        "GIS / Geospatial", "Regulatory Compliance Technology",
        "Disaster Resilience & Risk Reduction", "Climate Adaptation Technology",
    ],
    "technology_stack": [
        "Python", "FastAPI", "YOLO", "Vision Language Models", "LLMs",
        "Gemini", "AWS", "Cloudflare", "MongoDB", "PostgreSQL", "Docker",
        "REST APIs", "Vector Databases", "Real-time streaming", "ReportLab",
        "GIS/DEM pipelines",
    ],
    "products": [
        {
            "name": "AlgoVisual / AlgoOptic",
            "summary": (
                "Configurable Vision-AI surveillance platform that turns cameras into intelligent "
                "monitoring agents (YOLO + VLMs + LLM workflow engine) for industrial sites, "
                "warehouses, hospitals, and schools - PPE/safety, intrusion, fire & smoke, analytics."
            ),
        },
        {
            "name": "EmberCheck",
            "summary": (
                "NSW bushfire BAL (Bushfire Attack Level) screening tool built to AS 3959. Produces an "
                "indicative, conservative bushfire-risk read from address + photos + GIS/DEM data; an "
                "accredited assessor certifies and signs a PDF determination via a console. PropTech / "
                "disaster-resilience / regulatory-compliance SaaS."
            ),
        },
        {
            "name": "FloodCheck",
            "summary": (
                "NSW flood-risk screening platform (early stage) - automated property-level flood risk "
                "screening for buyers, assessors, and councils. PropTech / climate-adaptation / GIS."
            ),
        },
    ],
    "target_industries": [
        "Manufacturing", "Industrial facilities", "Mining", "Construction",
        "Warehousing & logistics", "Transportation", "Retail", "Education",
        "Smart cities", "Government", "Healthcare", "Commercial buildings",
        "Property / real estate", "Insurance & risk", "Local councils",
    ],
    "registration": {
        "australia_registered": False,
        "has_abn": False,
        "gst_registered": False,
        "australia_tax_registered": False,
        "australian_bank_account": False,
        "india_registered": False,
        "is_individual_applicant": False,
        "indigenous_owned": False,
        "registered_charity_nfp": False,
    },
    "previous_grants": [],
    "certifications": [],
}

# ---------------------------------------------------------------------------
# Layer A - deterministic blocker definitions
# ---------------------------------------------------------------------------
BLOCKER_RULES: List[Dict[str, Any]] = [
    {
        "keywords": ["abn", "australian business number", "registered business in australia",
                      "australian company", "acn"],
        "org_field": "has_abn",
        "reason": "Requires an ABN / Australian business registration (AlgoOrange is not registered in Australia yet)",
        "cap": 55,
        "tag": "abn_required",
    },
    {
        "keywords": ["gst", "goods and services tax"],
        "org_field": "gst_registered",
        "reason": "Requires GST registration (AlgoOrange is not GST registered)",
        "cap": 50,
        "tag": "gst_required",
    },
    {
        "keywords": ["australian bank account", "bank account in australia"],
        "org_field": "australian_bank_account",
        "reason": "Requires an Australian bank account",
        "cap": 50,
        "tag": "au_bank_required",
    },
    {
        "keywords": ["australian citizen", "permanent resident", "citizenship",
                      "resident of australia"],
        "org_field": "is_individual_applicant",
        "reason": "Requires Australian citizenship / residency",
        "cap": 30,
        "tag": "citizenship_required",
    },
    {
        "keywords": ["individuals only", "must be an individual", "sole trader only",
                      "not available to companies"],
        "org_field": "is_individual_applicant",
        "reason": "Open to individuals only; AlgoOrange is a company",
        "cap": 25,
        "tag": "individuals_only",
    },
    {
        "keywords": ["not-for-profit", "charity", "registered charity",
                      "incorporated association"],
        "org_field": "registered_charity_nfp",
        "reason": "Restricted to not-for-profits / charities",
        "cap": 25,
        "tag": "nfp_only",
    },
    {
        "keywords": ["farm", "farmer", "agricultur", "livestock", "crop", "forestry",
                      "fisher", "aquacultur", "childcare", "early childhood",
                      "mining lease", "tourism operator"],
        "org_field": None,
        "reason": "Grant is sector-locked to an unrelated industry (e.g. agriculture/childcare/forestry)",
        "cap": 30,
        "tag": "sector_locked",
    },
]


def _keyword_is_alternative(text: str, keyword: str) -> bool:
    """Check if keyword appears as one option among alternatives (e.g. 'X or a not-for-profit')."""
    idx = text.find(keyword)
    if idx == -1:
        return False
    context = text[max(0, idx - 80):idx + len(keyword) + 80]
    or_patterns = [" or a ", " or an ", " or be a ", " or be an ",
                   ", or ", " or not-for-profit", " or charity"]
    return any(p in context for p in or_patterns)


def run_layer_a(grant: Dict) -> Tuple[Optional[int], List[str], List[str]]:
    """Return (lowest_cap_or_None, list_of_reasons, list_of_tags)."""
    combined_text = " ".join([
        grant.get("check_if_you_can_apply") or "",
        grant.get("who_is_this_for") or "",
        grant.get("overview") or "",
    ]).lower()

    fired_reasons: List[str] = []
    fired_tags: List[str] = []
    lowest_cap: Optional[int] = None

    reg = ORGANIZATION_INFO["registration"]

    for rule in BLOCKER_RULES:
        hit_keywords = [kw for kw in rule["keywords"] if kw in combined_text]
        if not hit_keywords:
            continue

        if rule["org_field"] is not None:
            if reg.get(rule["org_field"], False):
                continue

        if rule["tag"] in ("nfp_only", "individuals_only", "citizenship_required"):
            all_alternative = all(
                _keyword_is_alternative(combined_text, kw) for kw in hit_keywords
            )
            if all_alternative:
                continue

        if rule["tag"] == "sector_locked":
            who_text = (grant.get("who_is_this_for") or "").lower()
            elig_text = (grant.get("check_if_you_can_apply") or "").lower()
            defining_audience = any(kw in who_text or kw in elig_text for kw in hit_keywords)
            if not defining_audience:
                continue

        fired_reasons.append(rule["reason"])
        fired_tags.append(rule["tag"])
        cap = rule["cap"]
        if lowest_cap is None or cap < lowest_cap:
            lowest_cap = cap

    return lowest_cap, fired_reasons, fired_tags


# ---------------------------------------------------------------------------
# LLM prompts
# ---------------------------------------------------------------------------
def build_system_prompt() -> str:
    org = ORGANIZATION_INFO
    products_text = "\n".join(
        f"  - {p['name']}: {p['summary']}" for p in org["products"]
    )
    return f"""You are an expert Australian business grants analyst scoring DOMAIN ALIGNMENT ONLY.

ORGANIZATION PROFILE:
Name: {org['name']}
Location: Sydney, Australia (operational base)
Type: {org['business_type']}
Team: {org['team']['size']} engineers - {org['team']['profile']}
Domains: {', '.join(org['industry_domains'])}
Tech stack: {', '.join(org['technology_stack'])}
Products:
{products_text}
Target industries: {', '.join(org['target_industries'])}

ASSUME: AlgoOrange is a fully registered Australian company with ABN, GST, and an Australian bank account. Registration, ABN, GST, citizenship, and bank account requirements are NOT your concern - a separate system handles those. NEVER mention ABN, GST, registration, or bank account in your analysis or let them affect the score.

YOUR TASK: Score how well the grant's PURPOSE, INDUSTRY, SECTOR, and GEOGRAPHIC FOCUS align with AlgoOrange's business.

SCORING GUIDE:
- 80-100%: Grant directly targets AI, technology, software, innovation, R&D, computer vision, PropTech, disaster resilience, digital transformation, or similar. AlgoOrange's products clearly match.
- 60-79%: Grant is for a broad category (small business support, exporters, energy efficiency, training, digital adoption, general NSW business) that an AI/tech company in Sydney could reasonably use.
- 40-59%: Some alignment (general business grants, R&D for any industry, NSW economic development) but grant's primary focus is a different sector. Still worth investigating.
- 20-39%: Grant targets a mostly unrelated sector (e.g. construction-specific, mining equipment, specific healthcare delivery) with only minor overlap.
- 0-19%: Grant is completely sector-locked to an unrelated industry (e.g. farming/livestock, childcare providers, forestry, fisheries) with zero relevance.

EXAMPLES to calibrate your scoring:
- "Funding for NSW small businesses to reduce energy costs" -> 65% (any NSW small business including tech, broad applicability)
- "R&D tax incentive for Australian companies" -> 85% (directly relevant to an R&D-heavy AI company)
- "Rebate for digital games production in NSW" -> 25% (AlgoOrange builds AI/PropTech, not games)
- "Funding for livestock transport operators" -> 0% (completely unrelated sector)
- "Disaster resilience infrastructure grants" -> 80% (directly aligns with EmberCheck/FloodCheck)
- "General export assistance for NSW companies" -> 70% (AlgoOrange could export AI/SaaS products)

WHAT TO INCLUDE in why_eligible:
- Geographic match (NSW/Sydney)
- Industry/technology relevance
- Product alignment (AlgoVisual, EmberCheck, FloodCheck)
- Company type match (accepts companies, SMEs, tech firms)
- Broad applicability ("open to all industries", "any NSW business")

WHAT TO INCLUDE in why_not_eligible:
- ONLY domain/sector mismatches (e.g. "Grant targets aged care providers, not tech companies")
- ONLY specific non-registration requirements that don't fit (e.g. "Must have 50+ employees")
- NEVER mention ABN, GST, Australian registration, or bank account.

For action_next_step: If the grant could be a fit, suggest concrete next steps like "Review detailed guidelines and prepare application" or "Contact grant provider to confirm tech companies can apply".

Respond with ONLY a single valid JSON object. No markdown, no code fences, no extra text.
Schema:
{{
  "eligibility_percent": <int 0-100>,
  "eligible": <bool>,
  "recommendation": "<STRONG CANDIDATE | WORTH PURSUING | NEEDS INVESTIGATION | NOT ELIGIBLE>",
  "why_eligible": ["short concrete bullet", "..."],
  "why_not_eligible": ["short concrete bullet", "..."],
  "geographic_fit": "one line",
  "industry_alignment": "one line",
  "international_company_eligible": "<Yes | No | Maybe>",
  "action_next_step": "one short actionable line",
  "reasoning": "2-3 sentences explaining the score"
}}"""


def build_user_prompt(grant: Dict) -> str:
    def trunc(val: Any, max_len: int = 2000) -> str:
        s = str(val) if val else "N/A"
        return s[:max_len] if len(s) > max_len else s

    return f"""Analyse this grant for AlgoOrange eligibility:

Title: {trunc(grant.get('title'))}
Status: {trunc(grant.get('status'))}
Close Date: {trunc(grant.get('closeDate'))}
Who is this for: {trunc(grant.get('who_is_this_for'))}
What do you get: {trunc(grant.get('what_do_you_get'))}
Overview: {trunc(grant.get('overview'))}
Eligibility criteria: {trunc(grant.get('check_if_you_can_apply'))}"""


# ---------------------------------------------------------------------------
# Groq call with retries
# ---------------------------------------------------------------------------
def call_groq(client: Groq, grant: Dict, model: str, logger: logging.Logger) -> Optional[Dict]:
    system_prompt = build_system_prompt()
    user_prompt = build_user_prompt(grant)

    for attempt in range(MAX_RETRIES + 1):
        try:
            resp = client.chat.completions.create(
                model=model,
                temperature=GROQ_TEMPERATURE,
                max_tokens=GROQ_MAX_TOKENS,
                timeout=REQUEST_TIMEOUT_S,
                response_format={"type": "json_object"},
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
            )
            raw_text = resp.choices[0].message.content
            return json.loads(raw_text)

        except json.JSONDecodeError:
            logger.warning("JSON parse failure on attempt %d for '%s'", attempt + 1, grant.get("title"))
        except Exception as e:
            err_str = str(e).lower()
            if "model" in err_str and ("not found" in err_str or "decommission" in err_str or "unknown" in err_str):
                logger.error(
                    "Model '%s' was rejected by Groq. "
                    "Check https://api.groq.com/openai/v1/models and update GROQ_MODEL.",
                    model,
                )
                print(
                    f"\nERROR: Model '{model}' was rejected by Groq. "
                    f"Check https://api.groq.com/openai/v1/models and update GROQ_MODEL."
                )
                return None
            logger.warning("API error on attempt %d for '%s': %s", attempt + 1, grant.get("title"), e)

        if attempt < MAX_RETRIES:
            delay = RETRY_BASE_DELAY_S * (2 ** attempt)
            logger.info("Retrying in %.1fs ...", delay)
            time.sleep(delay)

    return None


# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------
def reconcile(grant: Dict, llm_result: Optional[Dict], layer_a_cap: Optional[int],
              layer_a_reasons: List[str], layer_a_tags: List[str]) -> Dict:
    fallback = llm_result is None

    registration_leak_keywords = ["abn", "acn", "gst", "bank account", "not registered",
                                   "australian registration", "business number", "registered in australia"]

    if llm_result:
        model_score = int(llm_result.get("eligibility_percent", 50))
        why_eligible = list(llm_result.get("why_eligible", []))
        raw_why_not = list(llm_result.get("why_not_eligible", []))
        why_not_eligible = [
            r for r in raw_why_not
            if not any(kw in r.lower() for kw in registration_leak_keywords)
        ]
        recommendation = llm_result.get("recommendation", "NEEDS INVESTIGATION")
        action_next_step = llm_result.get("action_next_step", "")
        reasoning = llm_result.get("reasoning", "")
        geographic_fit = llm_result.get("geographic_fit", "")
        industry_alignment = llm_result.get("industry_alignment", "")
        international_eligible = llm_result.get("international_company_eligible", "")
    else:
        if layer_a_cap is not None:
            model_score = layer_a_cap
        else:
            model_score = 20
        why_eligible = []
        why_not_eligible = []
        recommendation = "NEEDS INVESTIGATION"
        action_next_step = "Manual review required - model was unavailable"
        reasoning = "Automatic fallback: model call failed; score based on deterministic blocker scan only."
        geographic_fit = ""
        industry_alignment = ""
        international_eligible = ""

    if layer_a_cap is not None:
        final_score = min(model_score, layer_a_cap)
    else:
        final_score = model_score

    existing_not = set(why_not_eligible)
    for reason in layer_a_reasons:
        if reason not in existing_not:
            why_not_eligible.append(reason)

    if grant.get("status") == "Open":
        open_bullet = "Grant is OPEN for applications"
        if open_bullet not in why_eligible:
            why_eligible.insert(0, open_bullet)

    eligible_yn = "YES" if final_score >= PURSUE_THRESHOLD else "NO"

    if final_score >= PURSUE_THRESHOLD and not fallback:
        if final_score >= TIER_LIKELY:
            recommendation = "STRONG CANDIDATE"
        elif recommendation == "NOT ELIGIBLE":
            recommendation = "WORTH PURSUING"

    if final_score < PURSUE_THRESHOLD and recommendation in ("STRONG CANDIDATE", "WORTH PURSUING"):
        recommendation = "NOT ELIGIBLE"

    if fallback:
        recommendation = "NEEDS INVESTIGATION (auto-fallback: model unavailable)"

    return {
        "title": grant.get("title"),
        "url": grant.get("url"),
        "status": grant.get("status"),
        "closeDate": grant.get("closeDate"),
        "final_score": final_score,
        "eligible": eligible_yn,
        "recommendation": recommendation,
        "why_eligible": why_eligible,
        "why_not_eligible": why_not_eligible,
        "action_next_step": action_next_step,
        "reasoning": reasoning,
        "geographic_fit": geographic_fit,
        "industry_alignment": industry_alignment,
        "international_company_eligible": international_eligible,
        "layer_a_cap": layer_a_cap,
        "layer_a_tags": layer_a_tags,
        "model_raw_score": model_score if llm_result else None,
        "fallback": fallback,
    }


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------
DARK_BLUE = PatternFill(start_color="FF1F3864", end_color="FF1F3864", fill_type="solid")
MEDIUM_BLUE = PatternFill(start_color="FF2E5496", end_color="FF2E5496", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
LIGHT_AMBER = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
LIGHT_RED = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
SECTION_FILL = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

WHITE_BOLD = Font(bold=True, color="FFFFFFFF", size=14)
WHITE_HEADER = Font(bold=True, color="FFFFFFFF", size=11)
GREEN_BOLD = Font(bold=True, color="FF006100")
RED_BOLD = Font(bold=True, color="FF9C0006")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

COL_WIDTHS = [8, 45, 10, 10, 12, 55, 55, 22, 40, 14, 40]
HEADERS = [
    "Grant ID", "Grant Title", "Status", "Eligible?", "Eligibility %",
    "Why Eligible", "Why Not Eligible", "Recommendation",
    "Action / Next Step", "Close Date", "URL",
]


def write_excel(results: List[Dict], total_input: int, analysed_count: int,
                closed_skipped: int, crawl_error_count: int, model_name: str,
                blocker_counts: Dict[str, int], filepath: str):
    wb = Workbook()

    # ---- Sheet 1: Eligibility Analysis ----
    ws = wb.active
    ws.title = "Eligibility Analysis"

    ws.merge_cells("A1:K1")
    cell_a1 = ws["A1"]
    cell_a1.value = "GRANT ELIGIBILITY ANALYSIS - ALGOORANGE"
    cell_a1.font = WHITE_BOLD
    cell_a1.fill = DARK_BLUE
    cell_a1.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:K2")
    now_str = datetime.now().strftime("%B %Y")
    cell_a2 = ws["A2"]
    cell_a2.value = f"Analysis Date: {now_str}  |  Total Grants Analysed: {analysed_count}  |  Model: {model_name}"
    cell_a2.font = Font(italic=True)
    cell_a2.alignment = Alignment(horizontal="center")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = WHITE_HEADER
        cell.fill = MEDIUM_BLUE
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[4].height = 28
    ws.auto_filter.ref = f"A4:K{4 + len(results)}"
    ws.freeze_panes = "A5"

    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, r in enumerate(results, start=5):
        grant_id = row_idx - 4
        score = r["final_score"]

        if score >= TIER_LIKELY:
            row_fill = LIGHT_GREEN
        elif score >= TIER_POSSIBLE:
            row_fill = LIGHT_AMBER
        else:
            row_fill = LIGHT_RED

        why_elig_str = " | ".join(r["why_eligible"]) if r["why_eligible"] else ""
        why_not_str = " | ".join(r["why_not_eligible"]) if r["why_not_eligible"] else ""

        row_data = [
            grant_id,
            r["title"],
            r["status"],
            r["eligible"],
            score,
            why_elig_str,
            why_not_str,
            r["recommendation"],
            r["action_next_step"],
            r["closeDate"] or "",
            r["url"] or "",
        ]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")

            if col_idx in (6, 7, 9):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        elig_cell = ws.cell(row=row_idx, column=4)
        elig_cell.font = GREEN_BOLD if r["eligible"] == "YES" else RED_BOLD

        url_cell = ws.cell(row=row_idx, column=11)
        if r.get("url"):
            url_cell.hyperlink = r["url"]
            url_cell.font = Font(color="FF0563C1", underline="single")

    # ---- Sheet 2: Summary ----
    ws2 = wb.create_sheet("Summary")

    ws2.merge_cells("A1:B1")
    cell_s1 = ws2["A1"]
    cell_s1.value = "ELIGIBILITY SUMMARY - ALGOORANGE"
    cell_s1.font = WHITE_BOLD
    cell_s1.fill = DARK_BLUE
    cell_s1.alignment = Alignment(horizontal="center", vertical="center")

    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 16

    likely = sum(1 for r in results if r["final_score"] >= TIER_LIKELY)
    possible = sum(1 for r in results if TIER_POSSIBLE <= r["final_score"] < TIER_LIKELY)
    unlikely = sum(1 for r in results if r["final_score"] < TIER_POSSIBLE)
    total_yes = sum(1 for r in results if r["eligible"] == "YES")

    metrics = [
        ("Metric", "Value", True, False),
        ("Total Grants in Input File", total_input, False, False),
        ("Open Grants (Analysed)", analysed_count, False, False),
        ("Closed / Skipped Grants", closed_skipped, False, False),
        ("Grants With Crawl Errors (skipped)", crawl_error_count, False, False),
        ("", "", False, False),
        ("ALGOORANGE ELIGIBILITY", "", True, True),
        ("Likely Eligible (>=70%)", likely, False, False),
        ("Possible Eligible (40-69%)", possible, False, False),
        ("Unlikely Eligible (<40%)", unlikely, False, False),
        ("Total Potentially Eligible (YES)", total_yes, True, False),
        ("", "", False, False),
        ("HARD BLOCKERS PRESENT", "", True, True),
        ("Grants requiring ABN/AU registration", blocker_counts.get("abn_required", 0), False, False),
        ("Grants requiring GST", blocker_counts.get("gst_required", 0), False, False),
        ("Grants requiring AU bank account", blocker_counts.get("au_bank_required", 0), False, False),
        ("", "", False, False),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), False, False),
        ("Model Used", model_name, False, False),
    ]

    for i, (label, value, bold, is_section) in enumerate(metrics, start=3):
        cell_a = ws2.cell(row=i, column=1, value=label)
        cell_b = ws2.cell(row=i, column=2, value=value)
        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER
        if bold:
            cell_a.font = Font(bold=True)
            cell_b.font = Font(bold=True)
        if is_section:
            cell_a.fill = SECTION_FILL
            cell_b.fill = SECTION_FILL

    try:
        wb.save(filepath)
    except PermissionError:
        alt = filepath.replace(".xlsx", "_v2.xlsx")
        logging.getLogger("grant_checker").warning(
            "File %s is locked, saving as %s", filepath, alt
        )
        wb.save(alt)
        filepath = alt

    return filepath


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def setup_logging(log_path: str) -> logging.Logger:
    logger = logging.getLogger("grant_checker")
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    return logger


def _strip_html(html_string: Optional[str]) -> Optional[str]:
    if not html_string:
        return None
    from bs4 import BeautifulSoup as _BS
    return " ".join(_BS(html_string, "html.parser").get_text(" ", strip=True).split())


def _ts_to_date(ts) -> Optional[str]:
    if not ts:
        return None
    try:
        return datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d")
    except (ValueError, OSError, TypeError):
        return None


def _normalize_coveo(result: Dict) -> Dict:
    """Convert a Coveo search result into the flat grant dict the checker expects."""
    raw = result.get("raw", {})
    return {
        "title": result.get("title") or raw.get("ctitle"),
        "url": "https://business.gov.au" + raw.get("curl", ""),
        "status": raw.get("cgs"),
        "closeDate": _ts_to_date(raw.get("closez32xdate")),
        "startDate": _ts_to_date(raw.get("startz32xdate")),
        "description": raw.get("csearchcarddescription"),
        "contact_phone": raw.get("ccontactphone"),
        "what_do_you_get": _strip_html(raw.get("whatz32xyouz32xget")),
        "who_is_this_for": _strip_html(raw.get("whoz32xthisz32xisz32xfor")),
        "overview": raw.get("fsearchz32xandz32xheaderz32xdescription28333"),
        "check_if_you_can_apply": None,
        "how_to_apply": None,
        "can_apply": False,
        "crawl_error": None,
    }


def main():
    parser = argparse.ArgumentParser(description="AlgoOrange Grant Eligibility Checker (Groq)")
    parser.add_argument("--input", default=INPUT_FILE_DEFAULT, help="Input grant JSON file")
    parser.add_argument("--limit", type=int, default=None, help="Analyse only first N grants")
    parser.add_argument("--all-statuses", action="store_true", help="Analyse all statuses, not just Open")
    parser.add_argument("--model", default=None, help="Override GROQ_MODEL")
    args = parser.parse_args()

    model = args.model or GROQ_MODEL
    only_open = ONLY_OPEN_GRANTS and not args.all_statuses

    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        print("ERROR: GROQ_API_KEY not found in environment variables.")
        print("Set it in your .env file or export it before running.")
        sys.exit(1)

    if not os.path.isfile(args.input):
        print(f"ERROR: Input file not found: {args.input}")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(OUTPUT_DIR, f"AlgoOrange_Grant_Eligibility_Analysis_{ts}.xlsx")
    raw_json_path = os.path.join(OUTPUT_DIR, f"grant_analysis_raw_{ts}.json")
    log_path = os.path.join(OUTPUT_DIR, f"run_log_{ts}.log")

    logger = setup_logging(log_path)

    with open(args.input, "r", encoding="utf-8") as f:
        raw_data = json.load(f)

    if isinstance(raw_data, dict) and "results" in raw_data:
        coveo_results = raw_data["results"]
        all_grants: List[Dict] = [_normalize_coveo(r) for r in coveo_results]
        logger.info("Detected Coveo API format (%d results)", len(all_grants))
    elif isinstance(raw_data, list):
        all_grants = raw_data
    else:
        print("ERROR: Unrecognized JSON structure (expected array or Coveo response)")
        sys.exit(1)

    total_input = len(all_grants)

    crawl_error_grants = [g for g in all_grants if isinstance(g, dict) and g.get("crawl_error")]
    crawl_error_count = len(crawl_error_grants)
    valid_grants = [g for g in all_grants if not g.get("crawl_error")]

    if only_open:
        grants_to_analyse = [g for g in valid_grants if g.get("status") == "Open"]
    else:
        grants_to_analyse = valid_grants

    if args.limit:
        grants_to_analyse = grants_to_analyse[:args.limit]

    analysed_count = len(grants_to_analyse)
    closed_skipped = total_input - analysed_count

    # Banner
    print("=" * 70)
    print("  ALGOORANGE - AUSTRALIAN GRANT ELIGIBILITY ANALYZER")
    print("=" * 70)
    print(f"  Organization : {ORGANIZATION_INFO['name']}")
    print(f"  Location     : {ORGANIZATION_INFO['location']}")
    print(f"  Model        : {model}")
    print(f"  Input file   : {args.input}")
    print(f"  Total grants : {total_input}")
    print(f"  Crawl errors : {crawl_error_count} (skipped)")
    print(f"  To analyse   : {analysed_count}")
    print("=" * 70)
    logger.info("Run started. Model=%s, Input=%s, Analysing=%d/%d grants",
                model, args.input, analysed_count, total_input)

    client = Groq(api_key=api_key)

    results: List[Dict] = []
    blocker_counts: Dict[str, int] = {}

    for idx, grant in enumerate(grants_to_analyse, start=1):
        title = (grant.get("title") or "Unknown")[:70]
        logger.info("[%d/%d] Analysing: %s", idx, analysed_count, title)

        layer_a_cap, layer_a_reasons, layer_a_tags = run_layer_a(grant)

        for tag in layer_a_tags:
            blocker_counts[tag] = blocker_counts.get(tag, 0) + 1

        llm_result = call_groq(client, grant, model, logger)

        final = reconcile(grant, llm_result, layer_a_cap, layer_a_reasons, layer_a_tags)
        results.append(final)

        status_icon = "YES" if final["eligible"] == "YES" else "NO"
        fb = " (FALLBACK)" if final["fallback"] else ""
        print(f"  [{idx}/{analysed_count}] {title} -> {final['final_score']}% {status_icon}{fb}")
        logger.info("  Result: %d%% %s %s%s",
                     final["final_score"], final["eligible"], final["recommendation"], fb)

        if idx < analysed_count:
            time.sleep(RATE_LIMIT_SLEEP_S)

    # Save raw JSON
    with open(raw_json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    logger.info("Raw results saved to %s", raw_json_path)

    # Generate Excel
    saved_xlsx = write_excel(
        results, total_input, analysed_count, closed_skipped,
        crawl_error_count, model, blocker_counts, excel_path,
    )
    logger.info("Excel saved to %s", saved_xlsx)

    # Final summary
    likely = sum(1 for r in results if r["final_score"] >= TIER_LIKELY)
    possible = sum(1 for r in results if TIER_POSSIBLE <= r["final_score"] < TIER_LIKELY)
    unlikely = sum(1 for r in results if r["final_score"] < TIER_POSSIBLE)
    total_yes = sum(1 for r in results if r["eligible"] == "YES")

    print("\n" + "=" * 70)
    print("  RESULTS SUMMARY")
    print("=" * 70)
    print(f"  Likely Eligible (>=70%)  : {likely}")
    print(f"  Possible Eligible (40-69%): {possible}")
    print(f"  Unlikely Eligible (<40%) : {unlikely}")
    print(f"  Total Potentially Eligible: {total_yes}")
    print("=" * 70)
    print(f"  Excel  : {saved_xlsx}")
    print(f"  JSON   : {raw_json_path}")
    print(f"  Log    : {log_path}")
    print("=" * 70)


if __name__ == "__main__":
    main()
